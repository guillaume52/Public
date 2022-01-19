# -*- coding: utf-8 -*-
"""
Created on Wed Jan  5 10:29:53 2022
 
@author: Guillaume Lombard
"""
 
"""
#OPTIMIZELY API references
#https://library.optimizely.com/docs/api/app/v2/index.html
"""
from easygui import *
import requests
import json
import re
import urllib
import pandas as pd
import numpy as np
import ast
import os
import time
import datetime
from openpyxl import load_workbook
import sys
#Token required for authentication
token='TOEKN HERE BETWEEN THE QUOTES (KEEP THE QUOTES)'
token='2:144e9gwG13pKp-jspUGgospelGm7CkfX7GdIWqFZ4rpze89_iytQ'
headers = {'Authorization': 'Bearer {}'.format(token)}
#create a dataframe that can de downloaded as a template
template=pd.DataFrame({'Audience name' : ["compulsory, this needs to match audience in optimizely"],
                        'AEM Audience': ["optional, this is only used for audience management - column must be present, data e.g. 21707387"],
                        'Audience Stream ID': ["compulsory, e.g 167"],
                        'Description': ["optional - describe the experiment"],
                        'Variation Name': ["compulsory, this is usually the VMO number and willbe used to name the variation"],
                        'Experiment Name': ["compulsory, can be anything as long as unique"],
                        'Experiment status': ["active, paused or archived - default to active"],
                        'Audience Description': ["optional - column must be present"],
                        'AEM image': ["compulsory,this is the banner location in AEM"],
                        'Extension_Id': ['Extension id can be found in the sheet called Extension'],
                        'AEM image side link': ["same as AEM Image, leave empty if not needed"],
                        'Extension_Id side link': ['Same as Extension_Id, leave empty if not required']})
 
#set global variable to make sure they within function, works but not the best, i did this as quick way of fixing some accessibility issue outside/inside function                 
global template_columns
global Pages_List
global Pages_id
global Audience_missing
global Audience_List
project_id=[]
#Collect the column names from the template, this will be used throughout to check that each time the script runs those columns exists
#this should allow for future modification e.g adding more columns. If removing columns, just make sure that the one that are being deleted aren't specifically mentioned further down (especially while merging)
template_columns=template.columns
 
#Create empty dataframe and delte them- ensure previous data isn't polluting new data set
df,QA_done,QA_df= pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
lst = [df,QA_done,QA_df]
del df,QA_done,QA_df # dfs still in list
del lst
 
#Fix issue when fetching data and item is missing to populate JSON - if no data is returned, one can specify what should be passed e.g. nan or ""
def try_or(func, default=None, expected_exc=(Exception,)):
    try:
        return func()
    except expected_exc:
        return default
 
#Open excel file
def openExcel():
    files=fileopenbox("Uploaad Excel file", "Optimizely", filetypes= "*.xlsx", multiple=False)
    xl=pd.ExcelFile(files)   
    sheetnames=choicebox("Which sheet contains the banners to upload?" , "Choose Sheet", sorted(xl.sheet_names))
 
    df=pd.read_excel(files,sheetnames, engine='openpyxl')
    try:
        df = df.convert_dtypes()
        df['Extension_Id side link']=df['Extension_Id side link'].str.replace(r'(\d+).\d+', r'\1')#convert columns to their likely types - this will fail in more than one type found in one colum e.g int and string, default is convert
        print("converted")
    except:
        df=df.astype(str)
        df['Extension_Id side link']=df['Extension_Id side link'].str.replace(r'(\d+).\d+', r'\1')
        print("cannot be converted automatically, convert to string")
     
    template_name= template_columns.tolist()
    df.drop_duplicates(subset=template_name, inplace=True) #remove potential duplicated row based on template columns
    #check that all expected columns are in the file the user is trying to open
    while True:
        if all(item in df.columns for item in template_name) is not True:
            print("column missing")
            msg = "Fix column and try again, if in doubt download template"
            title = "Column missing"
            if ccbox(msg, title):     # show a Continue/Cancel dialog
                pass  # user chose Continue
            else:  # user chose Cancel
                sys.exit(0)
            df=pd.read_excel(files, engine='openpyxl')
            return(df,files)              
             
        else:
            print("Columns are matching")
            return(df,files)
            break
#Open excel file           
#check if all columns required are in the file
#same as above, wrapped in a function to use further below
def columns_missing(df,files):
    while True:
        try:
            df.copy()[template_columns].isnull().values.any()
            print('success')
            break
        except KeyError:
            non_match = []
            for i in template_columns:
                if i not in df.columns:
                    non_match.append(i)
                    msgbox("Column missing(s): " + str(''.join(non_match)), "Columns missing", "Fix missing columns or upload will fail")
                    df=pd.read_excel(files, engine='openpyxl')           
#check if all columns required are in the file
##List audience avaible in Optimizely
def list_audiences(project_id,headers):
    global Audience_List
    global Audience_Value_parameter
    url="https://api.optimizely.com/v2/audiences"
    page=1
    per_page=100
    r_done = False
    r_total_json=[]
    Audience=[]   
    Audience_List=[]
    while True: #loop through all the pages until the last one          
        r=False
        params = {
        "per_page": 100,
        "page": page,
        "project_id" : project_id
        }
        if not r_done:  
            r = requests.get(url, params=params, headers=headers)
            if r.status_code == 200:
                print(r)
                r_total_json += r.json()
            r_done = 'Link' not in r.headers or 'rel=last' not in r.headers['Link']
        if r_done: # when both archived and non-archived have reached final page, exit while loop
            print(r_done)
            print(page)
            break
        page +=1 # get next page if no break
        print(page) 
  
    Audience=r_total_json 
     
    for P in Audience:
        audience_value=[]
        audience_value= try_or(lambda:str(ast.literal_eval(P['conditions'])[1][1][1]["value"]), default=str('nan'))
        Audience_List.append({
           'Audience name': P["name"],
           'Audience Value': audience_value,
           'Audience Description' : P["description"],
           'Audience Stream ID':re.sub(re.sub("[\d]+","",audience_value),"",audience_value),   #audience_value is read directly from optimizely        
            'Audience ID' : str(P['id']),
           'Project ID' : str(P['project_id']),
           'archived': P["archived"],
        })
    Audience_List=pd.DataFrame(Audience_List)
    audience_value=""   
    while audience_value=="" :
        try:   
            audience_value=Audience_List[Audience_List['Audience Value'].str.contains("hsbc_wpb")]['Audience Value'].iloc[0]   
            Audience_Value_parameter=re.sub("[\d]+","",audience_value)
        except IndexError:
            msgbox("Error missing Audience:you need to create at least one Tealium Audience manually in Optimizely for this script to work","Missing audience","Continue")
            print("done")
            break    
    return(Audience_List)
##List audience avaible in Optimizely
#List pages available in Optimizely
def list_pages(project_id,headers):
    url="https://api.optimizely.com/v2/pages"
    page=1
    per_page=100
    r_done = False
    r_total_json=[]
    Pages=[]   
    Pages_List=[]
    while True:          
        r=False
        params = {
        "per_page": per_page,
        "page": page,
        "project_id" : project_id
        }
        if not r_done:  
            r = requests.get(url, params=params, headers=headers)
            if r.status_code == 200:
                print(r)
                r_total_json += r.json()
            r_done = 'Link' not in r.headers or 'rel=last' not in r.headers['Link']
        if r_done: # when both archived and non-archived have reached final page, exit while loop
            print(r_done)
            print(page)
            break
        page +=1 # get next page if no break
        print(page) 
 
    Pages=r_total_json 
     
    for P in Pages:
        Pages_List.append({
           'Page name': P["name"],
           'Page URL': str(ast.literal_eval(P['conditions'])[1][1]['value']),
           'Page ID' : str(P['id']),
           'Project ID' : str(P['project_id']),
           'archived': P["archived"]
        })
        
    return(pd.DataFrame(Pages_List))
#List pages available in Optimizely
#List Experiments available
def list_experiments(project_id,headers,Campaign_id):
    page=1
    per_page=100
    r_done = False
    url="https://api.optimizely.com/v2/experiments"
    r_total_json=[]
    while True:
        r=False
        params = {
         "page" : page,  
        "per_page": per_page,
        "campaign_id" : Campaign_id,
        "project_id" : project_id
        }
     
        if not r_done:  
            r = requests.get(url, params=params, headers=headers)
            if r.status_code == 200:
                print(r)
                r_total_json += r.json()
            r_done = 'Link' not in r.headers or 'rel=last' not in r.headers['Link']
        if r_done: # when both archived and non-archived have reached final page, exit while loop
            print(r_done)
            print(page)
            break
        page +=1 # get next page if no break
        print(page)
    print(r_total_json)
     
    Experiment_df=[]
    for i in r_total_json:
        [s + "_Optimizely" for s in template.columns]
        Experiment_df.append({
                "Project ID" : i.get('project_id', ""),
                'Campaign id' : i.get('campaign_id', ""),
                'Created' : i.get('created', ""),
                'Experiment ID' : i.get('id', ""),
                'Experiment Name' : i.get('name', ""),
             'Description' : i.get('description', ""), 
                'page_id' :try_or(lambda: str(i['page_ids'][0]) if len(i['page_ids'])>= 1 else "", default=str('nan')) ,                   
             'Variation ID' :try_or(lambda:str(i['variations'][0]['variation_id']), default=str('nan')) ,
             'Variation Name' :try_or(lambda:str(i['variations'][0]['name']), default=str('nan')),   
             'AEM image' : try_or(lambda:str(i['variations'][0]['actions'][0]['changes'][0]['config']['aem_url']), default=str('nan')),
             'Extension_Id' : try_or(lambda: str(i['variations'][0]['actions'][0]['changes'][0]['extension_id']), default=str('nan')),            
             'AEM image side link' :try_or(lambda:str(i['variations'][0]['actions'][0]['changes'][1]['config']['aem_url']), default=str('nan')),
             'Extension_Id side link' :try_or(lambda: str(i['variations'][0]['actions'][0]['changes'][1]['extension_id']), default=str('nan')),
              'Audience ID' : try_or(lambda:str(ast.literal_eval(i['audience_conditions'])[1]['audience_id']), default=str('nan')),
                 'Experiment status' : i.get('status', ""),
                 'Variation status' :try_or(lambda:str( i['variations'][0]['status']), default=str('nan')) ,
              'Experiment URL' :'https://app.optimizely.com/v2/projects/{}/campaigns/{}/experiments/{}/variations/{}'.format(i.get('project_id', ""),i.get('campaign_id', ""),i.get('id', ""),try_or(lambda:str(i['variations'][0]['variation_id']), default=str('nan'))),             
              'Banner Preview Link' : try_or(lambda:str(i['variations'][0]['actions'][0]['share_link']), default=str('nan'))
                })
    Experiment_df=pd.DataFrame(Experiment_df)
    Experiment_df.columns
    return(Experiment_df)
#List Experiments avavible
#Prioritise banners
def prioritisation():
    QA_df=list_experiments(project_id,headers,Campaign_id)
    QA_df=QA_df[QA_df['Experiment status']!='archived']
    QA_df=QA_df.sort_values(by='Created', ascending=True).reset_index(drop=True) 
    Priorities_df=pd.merge(df[['Experiment Name', 'AEM image']].astype(str), QA_df[['Campaign id','Experiment Name', 'AEM image','Experiment ID']].astype(str), left_on=  ['Experiment Name', 'AEM image'], right_on= ['Experiment Name', 'AEM image'], how = 'left', suffixes=('', '_DROP')).filter(regex='^(?!.*_DROP)')
    Priorities_df['id']=QA_df['Experiment ID'].astype(str)
    Priorities_df['Match']= Priorities_df['id']==Priorities_df['Experiment ID']
    
    lists = []
    # make list of lists
    for i in (Priorities_df['Experiment ID']):
    # append list
        lists.append([i])
    # Display result
    print(lists)
 
    url="https://api.optimizely.com/v2/campaigns/"+str(Campaign_id)
    update_priorities={'experiment_priorities': lists}
    r = requests.patch(url, data=json.dumps(update_priorities), headers=headers)
    return(print(r.json()))   
#Prioritise banners  
#the below performs QA
def QA(Campaign_id): 
    global df
    global files
    QA_df=list_experiments(project_id,headers,Campaign_id)
    QA_df=QA_df[QA_df["Experiment status"] !="archived"]
    Audience_List=list_audiences(project_id,headers)
    Audience_List=Audience_List[Audience_List["archived"] ==False]
    Pages_List=list_pages(project_id,headers)
    Pages_List=Pages_List[Pages_List["archived"] ==False]
    #Left join Experiments list with Audience list
    QA_df=pd.merge(QA_df.astype(str), Audience_List.astype(str), left_on=  ["Project ID",'Audience ID'], right_on= ["Project ID",'Audience ID'], how = 'left', suffixes=('', '_OPTIMIZELY')).filter(regex='^(?!.*_OPTIMIZELY)')
    QA_df.columns
    #Left join QA_df with pages list
    QA_df=pd.merge(QA_df.astype(str), Pages_List.astype(str), left_on=  ["page_id"], right_on= ['Page ID'], how = 'left', suffixes=('', '_OPTIMIZELY')).filter(regex='^(?!.*_OPTIMIZELY)')
    QA_df.drop('Page ID', axis=1, inplace=True)
    #QA_df.drop_duplicates(subset=["Project ID",'Audience ID'], inplace=True)
    print(QA_df)
    #Drop duplicated rows based on Project ID and Audience ID - technically no drop should occuer as Optimizely doesnt allow audience id to be duplciated, else - you need to investigate
    Audience_List.drop_duplicates(subset=["Project ID",'Audience ID'], inplace=True)
    print(Audience_List)
    #Check that the excel has been uploaded and created as df, else, ask user to open it
    if 'df' in locals() or 'df' in globals():
        print("df exists")
    else:
        msgbox("Upload original file containing the banners details","Upload original file containing the banners details","Continue")
        df,files =openExcel()
     
    #The below list is usded to perform QA and columns that are compulosry to create a banner (i have exlcuded all columns containing "description" becasue they arent critical to create banners)
    merge_list=['Audience name', 'Audience Stream ID', 'Variation Name', 'Experiment Name', 'AEM image']
    print(df)
    #This will be used to allow duplciated columns in the final QA excel sheet, so a visual comparison can be done between the orignal file and the data contained in Optimizely
    QA_df.columns=[s + "_Optimizely" for s in QA_df.columns]
    #Left join orignal Excel and the QA dataframe
    QA_done=pd.merge(df.astype(str),QA_df.astype(str),left_on=merge_list, right_on=[s + "_Optimizely" for s in merge_list] , how = 'left')
 
    #hte below performs comparison between coulmns and value found in optimizely vs excel dataframe                   
    for QAindex, QA_row in QA_done[0:len(QA_done)].iterrows():
    #for QAindex, QA_row in QA_done[0:1].iterrows():
        print(QAindex)
        merge_list_append=[]
        QA_missing_items=[]
        upload_status=""
        if pd.isna(QA_row['page_id_Optimizely']):
            for j in range(0,len(merge_list)):
                #print(j,merge_list[0:j])
                merge_list_append.append(j)
                print(merge_list_append)
                QA_check=[]        
                QA_check=pd.merge(df[df['Experiment Name']==QA_row['Experiment Name']].astype(str),QA_df.astype(str),left_on=merge_list, right_on=[s + "_Optimizely" for s in merge_list] , how = 'left')           
                
                print(QA_check['page_id_Optimizely'])
 
                #check if page_id_Optimizely exist, I could have picked up any other columns, if it contains nana, then it means the left join has failed to return a match and something is missing
                if pd.isna(QA_check['page_id_Optimizely']).tolist()==True: 
                    #This is an attempt to identify which column is responsible for the failed left join
                    QA_missing_items.append(merge_list[j])
                    print(QA_missing_items)
                    #originally I planned to identify which column can be matched on and which can't, then run a left join on each column, but this prove to be overly complicated and had no time to complete.
                    #the below doesnt really serve any purpose but I left it there just in case someone wnats to revisit this later on
                    #Depending on the issue, QA_missing_items will contain info pointing int he right direction, the bare minimm is you get nana in row missing info to indicate that something hasnt matched
                    merge_list_append.remove(j)
                    #'Upload Status' column only exist if the script was used to upload a "New Campaign", if the column doesnt exist, I need return en empty sstring so the script doesnt break due to columnnot found
                    try:
                        upload_status=''.join(QA_done['Upload Status'])
                    except:
                        upload_status= ""
            if  QA_missing_items:                
                QA_done.loc[QA_done['Experiment Name'] ==QA_row['Experiment Name'], 'QA Status'] = upload_status + ". Issue likely in:" + ','.join(QA_missing_items)
            else:
                QA_done.loc[QA_done['Experiment Name'] ==QA_row['Experiment Name'], 'QA Status'] = "Experiment Likely not uploaded.If it has been uploaded, then it is likely that on of the compulsory filed is not matching, start by checking that the setup of the audience in optimizely is mathcing the Audience ID"
                 
        else:
             QA_done.loc[QA_done['Experiment Name'] ==QA_row['Experiment Name'], 'QA Status']="It's looking good however check if any rows contains nan, this will indicate that something is missing "
    #Check if the number of experiment in the excel is matching the one in Optimizely 
    if len(df.index) > len(QA_df.index):
        msgbox("Missing Experiment","Some Epxeriments are missing in optimizely - See QA file for ore info","Continue")
        print("Some Epxeriments are missing in optimizely - See QA file for ore info")
    elif len(df.index) < len(QA_df.index):
        msgbox("More Experiment in Optimizely than expected - See QA file for ore info","Too many Experiment","Continue")
        print("More Experiment in Optimizely than expected - See QA file for ore info")
    else:
        msgbox("number of experiments seems correct, check file for further info","Experiment","Continue")
        print("number of experiments seems correct, check file for further info") 
     
    #create the path and naming convention for the output file
    path = os.path.join(os.path.dirname(files), os.path.basename(re.sub(".xlsx","_QA_completed_"+re.sub(":","-",str(datetime.datetime.now()) ) +".xlsx",files)))       
    writer = pd.ExcelWriter(path, engine='openpyxl')
    # copy existing sheets
    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
    # write out the new sheet
    #Remove <NA> reated when converting df to string
    df=df.replace('<NA>', '')
    df.to_excel(writer,sheet_name="Original", index=False, header=True,startrow=0, columns=df.columns)
    QA_done.to_excel(writer,sheet_name="QA", index=False, header=True,startrow=0, columns=QA_done.columns)
    QA_df.to_excel(writer,sheet_name="Experiment in Optimizely", index=False, header=True,startrow=0, columns=QA_df.columns)
    Audience_List.to_excel(writer,sheet_name="Complete Optimizely Audience", index=False, header=True,startrow=0, columns=Audience_List.columns)
    writer.close()
    return(df,QA_done) 
       
 
 
 
##List Projects
def list_projects():
    params = {
    "per_page": 100
    }
     
    url="https://api.optimizely.com/v2/projects"
    r = requests.get(url, params=params, headers=headers)
    Projects=[]
    Project_List = []
    Projects=r.json()
    for P in Projects:
        Project_List.append({
           'name': P["name"],
           'id': P["id"],
           'description': P["description"]
        })
     
    Project_List=pd.DataFrame(Project_List)
    return(Project_List)
 

 
    
##List Campaigns
def List_campaigns(project_id):
    global Campaigns   
    page=1
    per_page=100
    r_done = False
    url="https://api.optimizely.com/v2/campaigns"
    r_total_json=[]
    while True:
        r=False
        params = {
         "page" : page,  
        "per_page": per_page,
        "project_id" : project_id
        }
     
        if not r_done:  
            r = requests.get(url, params=params, headers=headers)
            if r.status_code == 200:
                print(r)
                r_total_json += r.json()
            r_done = 'Link' not in r.headers or 'rel=last' not in r.headers['Link']
        if r_done: # when both archived and non-archived have reached final page, exit while loop
            print(r_done)
            print(page)
            break
        page +=1 # get next page if no break
        print(page)
    print(r_total_json)
    Campaigns=[]
    Campaigns_List=[]
    Campaigns=r.json()
     
    for P in Campaigns:
        print( P["name"], P['type'])
        print(P['name'],P['id'],P['type'],P['status'])
        if 'archived' not in P['status']:
            print(P['status'])
            Campaigns_List.append({
               'name': P["name"],
               'id' : P['id'],
               'type' : str(P['type']),
               'description': P["description"],
               'created' : P["created"],
               'status': P['status'],
               'type' : P['type'],
               'metrics' : P['metrics'],
               'last_modified' : P['last_modified']
            })
                 
    return(pd.DataFrame(Campaigns_List))
def list_extensions():
    url="https://api.optimizely.com/v2/extensions"
    params = {
    "per_page": 100,
    'project_id' : project_id
    }
    r = requests.get(url, params=params, headers=headers)
    Extensions=[]
    Extensions_List=[]
    Extensions=r.json()
     
    for P in Extensions:
        #print(P)
        print(P['name'],P['id'],P["description"])
        Extensions_List.append({
           'name': P["name"],
           'id' : P['id'],
           'description': P["description"]
        })
     
    return(pd.DataFrame(Extensions_List))
 
def upload_experiments(df):
        global Pages_id
        #the below values are defaulted to go through each rows of the excel file, they can be amended foe debug purpse if yo want to e.g upload only a handful of banner instead of the all list.
        #the start row is 0, last row the number of row minus 1.those value can be amended in that range
        start_index=0
        end_index=len(df.index)
        #end_index=2
        for index, row in df[start_index:end_index].iterrows():
            print(index)
            #I set all the variables by default to False instead of empty, so I can detect an issue and fail the upload if any
            #if Empty it would be hard to differntiate between an empty excel cell left intentionally blank or by mistake
            description=False
            experiment=False
            Aem_image=False
            Extensions_id=False
            Aem_image_side_link=False
            Extensions_id_side_link=False
            Audience_id=False
            Audience_Name=False
            image_dict=[]
            description=row["Description"]
            experiment=row["Experiment Name"]
            Aem_image=row["AEM image"]
            Extensions_id=str((row['Extension_Id']))
            Aem_image_side_link=row['AEM image side link']
            #the below is fiddly and trying to juggle between data type and understand if there is a value in the column
            #I m trying to detect nan or int, but it changes depending if it s a brand new upload from excel of a conitnued upload where data has bto be checked against Optimizely data
            #if  (row['Extension_Id side link']).isdigit():
            if  str(row['Extension_Id side link']).isdigit():    
                Extensions_id_side_link= str(int(row['Extension_Id side link']))         
                image_dict = [{"async": False,
                           "config": {"aem_url":Aem_image },
                           "extension_id": Extensions_id,
                           "type": "extension"},
                        {"async": False,
                         "config": {"aem_url": Aem_image_side_link},
                         "extension_id": Extensions_id_side_link,
                         "type": "extension"}
                        ] 
            elif str(row['Extension_Id side link']) == '<NA>':
                image_dict = [{"async": False,
                           "config": {"aem_url":Aem_image },
                           "extension_id": Extensions_id,
                           "type": "extension"}
                              ]
            else:
                image_dict = [{"async": False,
                           "config": {"aem_url":Aem_image },
                           "extension_id": Extensions_id,
                           "type": "extension"}
                              ]
                
             
            for indexation, value in Audience_List[0:len(Audience_List)].iterrows():  
                if 'archived':
                    if str(value['Audience name']) in str(row["Audience name"]):
                        print(value['Audience name'],value['Audience ID'],row["Audience name"])
                        Audience_id=value['Audience ID']
                        Audience_URL="https://app.optimizely.com/v2/projects/{}/audiences/{}/".format(project_id,Audience_id)
             
            Variation_Name=row["Variation Name"]
            print(Pages_id)
            experiment_payload={'allocation_policy': 'manual',
              'audience_conditions': '["and", {"audience_id": %s }]' % (str(Audience_id)),
              'campaign_id': Campaign_id,
              'changes': [],
              'description': str(description),
              'is_classic': False,
              'metrics': [],
              'name': str(experiment),
              'page_ids': [Pages_id],
              'project_id': project_id,
              'status': 'not_started',
              'traffic_allocation': 10000,
              'type': 'personalization',
              'variations': [{'actions': [{'changes':image_dict,
                  'page_id': Pages_id}],
                'archived': False,
                'name': str(Variation_Name),
                'status': 'active',
                'weight': 10000}]}
             
             
         
            element_missing=[]
             
            try:
                #the below if, allows to keep using this whole function between "New campaigns", "continue upload" and "update campaign"
                #Checking for _merge, indicates that we are updating an existing campaign, that column wont exist if you are uploading a brand new campaign
                if '_merge' and 'Project ID_y' and 'Experiment ID_y'  in df:
                     
                    if 'both' in row['_merge']:
                        df.loc[df['Experiment Name']==experiment,'Upload Status']="Experiment Not Updated - exist already and match file"
                        print("Experiment Not Updated - exist already and match file") 
                    elif row['Experiment ID_y'].isdigit() and 'both' not in row['_merge']: # Optimizely doesnt allow updating existing Perzonalisation experiment, only AB test, so the below wont work
                    #but I mleaving it there so the sscript doesnt break
                        #url = 'https://api.optimizely.com/v2/experiments/'+row['Experiment ID_y']
                        #r = requests.patch(url, data=json.dumps(experiment_payload), headers=headers)
                        print("Experience updated")
                    else:
                        #the below will create a new experiment in  existing campaign, hence whiy the the above is still needed
                        url = 'https://api.optimizely.com/v2/experiments'                       
                        r = requests.post(url, data=json.dumps(experiment_payload), headers=headers)
                        print("Experiment created")
                else:
                    url = 'https://api.optimizely.com/v2/experiments'
                    
                    r = requests.post(url, data=json.dumps(experiment_payload), headers=headers)
                    print("Experiment Created")
 
                    r.raise_for_status()
                    Experiments=r.json()
                    #This collets all the info about the newly created experience and add them to the df dataframe
                    df.loc[df['Experiment Name']==experiment,'Project ID']=str(project_id)
                    df.loc[df['Experiment Name']==experiment,'Campaign id']=str(Campaign_id)       
                    df.loc[df['Experiment Name']==experiment,'Experiment ID']=str(Experiments['id'] )
                    df.loc[df['Experiment Name']==experiment,'Variation ID']=str(Experiments['variations'][0]['variation_id'])
                    df.loc[df['Experiment Name']==experiment,'Experiment URL']='https://app.optimizely.com/v2/projects/{}/campaigns/{}/experiments/{}/variations/{}'.format(project_id,Campaign_id,Experiments["id"],Experiments['variations'][0]['variation_id'])
                    df.loc[df['Experiment Name']==experiment,'Banner Preview Link']=Experiments['variations'][0]['actions'][0]['share_link']
                    df.loc[df['Experiment Name']==experiment,'Audience ID']=str(Audience_id)
                    df.loc[df['Experiment Name']==experiment,'Audience link']=str(Audience_URL)
          
                    #check that the info returned exist (is not False, "" or empty is still considered to be a value)
                    if description and experiment and Aem_image and Audience_id:
                        print("Success")
                        df.loc[df['Experiment Name']==experiment,'Upload Status']="Experiment created succesfully"
                    if not description or not experiment or not Aem_image or not Audience_id:
                        for index, i in enumerate([description,experiment,Aem_image,Audience_id]):
                             if "False" in str([i]) :
                                element_missing +=["description","experiment","Aem_image","Audience_id"][index]
                                print('Experiment created but value missing on '+' '.join(element_missing))
                                df.loc[df['Experiment Name']==experiment,'Upload Status']='Experiment created but value missing on '+' '.join(element_missing)
            #except will capture any failed upload which could occur
            #every now and then you will get a 504, this is a server timeout from Optimizely's side, nothing we can do about it. owever the script wil handle them and continue as sooon as the server allow connection again
            #400 usually an issue that is created by assing the wrong type of information in the API (see dtypes)                  
            except requests.exceptions.HTTPError as e:
                df.loc[df['Experiment Name']==experiment,'Upload Status']="Error: " + str(e) +"."
                print("Experiment not created due to error: " + str(e))
                #experiment_payload  
             
            #Experiments
            print("End of loop " + str(index))
         
 
        #Run QA when the loop is completed
        df,QA_done=QA(Campaign_id)
        return(df,QA_done)
        #else:
           # return(df)
def audience_missing(Audience_List):  
    #this script checks that the audiences we are trying to use actually exists in optimizely, if not then it will attempt to create them.
    #in case this fails, upload the audience manaully, and make sure the excel file refelcts all the info on them so the file matches
     
    global Audience_missing
    Audience_missing=[]
    #left join audience infor from df and Audience info downloaded from the API
    Audience_missing=pd.merge(df["Audience name"],Audience_List, on ="Audience name",how='left')
    Audience_missing=Audience_missing[Audience_missing['archived'] != True]
 
    #check for discrpeancy betwen the 2 files based on Audience value
    '''
    I was made aware a few days before I left the company, that there might be a need to blend the tealium audience with Optimizely attributes - and dont have the time to fix it
    e.g tealium audience + Optimizely device -== 'mobile'
    the below script will likely fails as it won't be able to match the df with what exists in optimizely.
    a quick fix would be to change 'Audience Value' in the below with "Audience name" so the script only checks that the namingc ovnetionmatch rather than the content of the audience
    Then you would need to QA the audience manually tomake sure they contain al the information required
    '''
    while Audience_missing['Audience Value'].isna().any().tolist() or Audience_missing['Audience Value'].isnull().any().tolist() or (Audience_missing['Audience Value'] =="").any().tolist():
         
        print("Missing Audience")
        Audience_missing.columns
        Audience_missing=pd.merge(df[["Audience name",'Audience Stream ID','Audience Description','AEM Audience']],Audience_List, on ="Audience name",how='left', suffixes=('', '_DROP')).filter(regex='^(?!.*_DROP)')
        Audience_missing=Audience_missing.drop_duplicates(subset=['Audience name'])
        Audience_missing=Audience_missing[(Audience_missing['Audience Value'].isna()) |(Audience_missing['Audience Value'].isnull())|((Audience_missing['Audience Value'] ==""))]
        for index, empty in Audience_missing[0:len(Audience_missing)].iterrows():
        #for index, empty in Audience_missing[0:2].iterrows():
            #print(index)
            #in order for this to work one audience coming from tealium must be present in Optimizely, becasue the parameterused for each market differs
            #e.g HSBC UK = hsbc_wpb-stream-uk, FD =hsbc_wpb-stream-uk-fd_ etc...
            #there is no way I can read this value off so one at least MUST be manually created in Optimizely so I can read it
            if empty["Audience Value"]=="":
                audience_id_update=[]
                update_audience=[]
                audience_value_update=[]
                description_update=[]
                name_update=[]
                name_update=empty["Audience name"]
                audience_id_update=empty['Audience ID']
                audience_value_update=Audience_Value_parameter+"".join(Audience_missing[Audience_missing['Audience ID']==audience_id_update]['Audience Stream ID'].astype(str).values.tolist())
                description_update= " ".join(Audience_missing[Audience_missing['Audience name']==name_update]["Audience Description"].astype(str).values.tolist()) + " | " + " ".join(Audience_missing[Audience_missing['Audience name']==name_update]['AEM Audience'].astype(str).values.tolist())                               
                update_audience={"conditions" : '["and", ["or", ["or", {"name": "tealium.audiences", "type": "third_party_dimension", "value":' +'\"'+ audience_value_update+'\"' + '}]]]',
                        "description": description_update,
                        "name": name_update
         
                }
                print(audience_id_update)
                print(update_audience)
                url="https://api.optimizely.com/v2/audiences/{}".format(str(audience_id_update))
                r = requests.patch(url, data=json.dumps(update_audience), headers=headers)
                r.json()
            elif np.isnan(empty["Audience Value"]):
                upload_audience=[]
                audience_value_upload=[]
                description_upload=[]
                name_upload=[]
                name_upload=empty["Audience name"]
                audience_value_upload=Audience_Value_parameter+"".join(Audience_missing[Audience_missing['Audience name']==name_upload]['Audience Stream ID'].astype(str).values.tolist())
                description_upload=" ".join(Audience_missing[Audience_missing['Audience name']==name_upload]["Audience Description"].astype(str).values.tolist()) + " | " +   " ".join(Audience_missing[Audience_missing['Audience name']==name_upload]['AEM Audience'].astype(str).values.tolist())
                 
                upload_audience={"conditions" : '["and", ["or", ["or", {"name": "tealium.audiences", "type": "third_party_dimension", "value":' +'\"'+ audience_value_upload+'\"' + '}]]]',
                    "description": description_upload,
                    "name": name_upload,
                    "project_id" : project_id
                }
                print(upload_audience)
                url="https://api.optimizely.com/v2/audiences"
                r = requests.post(url, data=json.dumps(upload_audience), headers=headers)
                r.json()
       
        Audience_List = list_audiences(project_id,headers)
        Audience_missing=pd.merge(df["Audience name"],Audience_List, on ="Audience name",how='left')
        Audience_missing=Audience_missing[Audience_missing['archived'] != True]
        Audience_missing=pd.merge(df[["Audience name",'Audience Stream ID','Audience Description','AEM Audience']],Audience_List, on ="Audience name",how='left', suffixes=('', '_DROP')).filter(regex='^(?!.*_DROP)')
        Audience_missing=Audience_missing.drop_duplicates(subset=['Audience name'])
        return(Audience_missing[(Audience_missing['Audience Value'].isna()) |(Audience_missing['Audience Value'].isnull())|((Audience_missing['Audience Value'] ==""))])          
        print("done")
        #return(pd.DataFrame(Audience_missing))
     
def pull_QA_List():
    global QA_df
    QA_df=list_experiments(project_id,headers,Campaign_id)
    QA_df=QA_df[QA_df["Experiment status"] !="archived"]
    Audience_List=list_audiences(project_id,headers)
    Audience_List=Audience_List[Audience_List["archived"] ==False]
    Pages_List=list_pages(project_id,headers)
    Pages_List=Pages_List[Pages_List["archived"] ==False]
    QA_df=pd.merge(QA_df.astype(str), Audience_List.astype(str), left_on=  ["Project ID",'Audience ID'], right_on= ["Project ID",'Audience ID'], how = 'left', suffixes=('', '_DROP')).filter(regex='^(?!.*_DROP)')
    QA_df=pd.merge(QA_df.astype(str), Pages_List.astype(str), left_on=  ["page_id"], right_on= ['Page ID'], how = 'left', suffixes=('', '_DROP')).filter(regex='^(?!.*_DROP)')
    QA_df.drop('Page ID', axis=1, inplace=True)
 
        #QA_df.drop_duplicates(subset=["Project ID",'Audience ID'], inplace=True)
         
def duplicate_experiments():
    global Pages_id
    Pages_List=list_pages(project_id,headers)
    Pages_List=Pages_List[Pages_List['archived'] != True]
    Pages_id=int(Pages_List.loc[Pages_List["Page name"]==choicebox("Select the page where the campaign runs on."  , "Optimizely Uploader",  sorted(Pages_List["Page name"]) ) ,'Page ID'].item())
     
    dup=[]
    pull_QA_List()   
    duplicate_exp=[]
    duplicate_exp=QA_df.duplicated(subset=['Experiment Name', 'Variation Name','AEM image', 'Audience ID','Audience name', 'Audience Value', 'Audience Stream ID', 'archived'], keep=False)
    duplicated_exp=[]
    duplicated_exp=QA_df[QA_df.duplicated(subset=['Experiment Name', 'Variation Name','AEM image', 'Audience ID','Audience name', 'Audience Value', 'Audience Stream ID', 'archived'], keep=False)]
    duplication_choice=[]
    if (duplicate_exp ==True).any().tolist():
            dup="duplicates"
            button_list=["Delete the oldest","Abort","Delete the newest","Continue as it is"]
            duplication_choice=buttonbox("You have duplicated expepriments: "+' ,###, '.join(list(set(duplicated_exp['Experiment Name'].tolist()))), "Duplicated experiments", button_list)
            print("continue")
    else:
        dup="no duplicates"
    if duplication_choice =="Delete the oldest":
        duplicated_exp=duplicated_exp.sort_values(by=['Experiment Name','Created'], ascending=True)[['Experiment Name','Created','Experiment ID']]
        duplicated_exp=duplicated_exp[duplicated_exp.duplicated(subset=['Experiment Name'], keep='last')]
        for del_exp in duplicated_exp['Experiment ID']:
            print(del_exp)
            url='https://api.optimizely.com/v2/experiments/'+del_exp
            r=requests.delete(url,headers=headers)
            r
     
    elif duplication_choice == "Abort":
        sys.exit(0)
    elif duplication_choice == "Delete the newest":
        dup="duplicates"
        duplicated_exp=duplicated_exp.sort_values(by=['Experiment Name','Created'], ascending=False)[['Experiment Name','Created','Experiment ID']]
        duplicated_exp=duplicated_exp[duplicated_exp.duplicated(subset=['Experiment Name'], keep='last')]
        for del_exp in duplicated_exp['Experiment ID']:
            print(del_exp)
            url='https://api.optimizely.com/v2/experiments/'+del_exp
            r=requests.delete(url,headers=headers)
            r
         
    elif duplication_choice == "Continue as it is":
        dup="duplicates"
        print("continue")
     
    return(dup)
 
Project_List=list_projects()      
Menu_selection=[]
#sorted(Campaigns_List["name"])
Menu_selection=choicebox("Select the campaign you want to work on"  , "Optimizely Uploader", sorted(["Download experiment details","Update experiments - can only add new ones and prioritise","Continue Upload","New Campaign","Download template","QA","Check AEM link"]))
 
 
if Menu_selection =="New Campaign":
    #GUI to select project/project id
    project_id=[]
    project_id=Project_List.loc[Project_List['name']==choicebox("Select the project you want to work on.Warning if duplicate name exist, only first in the list willbe selected"  , "Optimizely Uploader", sorted(Project_List["name"]) ) ,'id'].item()
    ##List projects
    #List audience
    Audience_List=list_audiences(project_id,headers)       
    ###List Audience
    df,files=openExcel()    
    columns_missing(df,files)               
    Audience_List = list_audiences(project_id,headers)        
    audience_missing(Audience_List)       
    #not necessary anymore since implementing side link, however extension id need to be in the excel
    #Extensions_List=list_extensions()
    #Extensions_id=Extensions_List.loc[Extensions_List['name']==choicebox("Select the extension where the banner will be implemnted on"  , "Optimizely Uploader",  sorted(Extensions_List["name"]) ) ,'id'].item()
 
    Pages_List=list_pages(project_id,headers)
    Pages_List=Pages_List[Pages_List['archived'] != True]
    Pages_id=int(Pages_List.loc[Pages_List["Page name"]==choicebox("Select the page where the campaign runs on."  , "Optimizely Uploader",  sorted(Pages_List["Page name"]) ) ,'Page ID'].item())
     
    ##create campaign
    Campaign_name=[]   
    Campaign_name_custom = enterbox("Campaign name","Optimizely Uploader") 
    Campaigns_List=List_campaigns(project_id)
    while True:
        if Campaign_name_custom in [x for x in  Campaigns_List['name'] if x in Campaign_name_custom]:
            Campaign_name_custom = enterbox("Campaign name exists already chosse another one","Optimizely Uploader")
        else:
            Campaign_name = Campaign_name_custom
            break                       
    #create new campaign   
    payload={
      "description": Campaign_name ,
      "holdback": 0,
      "name": Campaign_name,
      "project_id": project_id,
      "type": "personalization",
      'page_ids': [Pages_id]
      }
 
    url="https://api.optimizely.com/v2/campaigns"
    r = requests.post(url, data=json.dumps(payload), headers=headers)
    Campaign_created=[]
    Campaign_created=r.json()
    Campaign_id=Campaign_created["id"]
    #create new campaign
     
    start = time.time()
    df.columns
    #upload the banners
    upload_experiments(df)       
    end = time.time()
    print(end - start)
    print(str(datetime. timedelta(seconds=end - start)))
 
elif Menu_selection == "Download template":
     
    project_id=[]
    project_id=Project_List.loc[Project_List['name']==choicebox("Select the project you want to work on.Warning if duplicate name exist, only first in the list willbe selected"  , "Optimizely Uploader", sorted(Project_List["name"]) ) ,'id'].item()
    project_name=Project_List[Project_List['id']==project_id]['name'].tolist()[0]
    Audience_List=list_audiences(project_id,headers)       
    Pages_List=list_pages(project_id,headers)
    Campaigns_List=List_campaigns(project_id)
    Campaigns_List=Campaigns_List.sort_values(by='created', ascending=False)
    Extensions_List=list_extensions() 
    Extensions_List.sort_values(by='name', ascending=True)
    directory = diropenbox('Select location for template')
    writer = pd.ExcelWriter(os.path.join(directory,project_name+"_Optimizely_template.xlsx"), engine='openpyxl')
    template.to_excel(writer,sheet_name="Template", index=False, header=True,startrow=0, columns=template.columns)
    Audience_List.to_excel(writer,sheet_name="Audience in Optimizely", index=False, header=True,startrow=0, columns=Audience_List.columns)
    Campaigns_List.to_excel(writer,sheet_name="Campaigns in Optimizely", index=False, header=True,startrow=0, columns=Campaigns_List.columns)
    Extensions_List.to_excel(writer,sheet_name="Exensions in Optimizely", index=False, header=True,startrow=0, columns=Extensions_List.columns)
    Pages_List.to_excel(writer,sheet_name="Pages in Optimizely", index=False, header=True,startrow=0, columns=Pages_List.columns)
    writer.close()
    msgbox("Template can be found under:" + os.path.join(directory,project_name+"_Optimizely_template.xlsx"), "File ready", "OK")
 
elif Menu_selection =="QA":
    df,files=openExcel()
    #GUI to select project/project id
    project_id=[]
    project_id=Project_List.loc[Project_List['name']==choicebox("Select the project you want to work on.Warning if duplicate name exist, only first in the list willbe selected"  , "Optimizely Uploader", sorted(Project_List["name"]) ) ,'id'].item()
    ##List projects
    #List audience
    Audience_List=list_audiences(project_id,headers)       
    ###List Audience
    Campaigns_List=[]
    Campaigns_List=List_campaigns(project_id)
    #this sort the list of campaing by descending order based on time of creation - change value below e.g by='name', ascending=True) for ascending alpahbetical order on campaign name
    Campaigns_List=Campaigns_List.sort_values(by='created', ascending=False, na_position='first')
    Menu_selection=choicebox("Select the campaign you want to work on"  , "Optimizely Uploader", (Campaigns_List["name"]))
    Campaign_id=[]
    Campaign_id=Campaigns_List.loc[Campaigns_List['name']==Menu_selection ,'id'].item()
    df,QA_done=QA(Campaign_id)
    #del df,QA_done,QA_df
 
elif Menu_selection =="Check AEM link":
    df,files=openExcel()
    URL_column=choicebox("Select the column containing the URLs (MUST include http - not just AEM link)"  , "URL checker",sorted(df.columns) )   
    df['Valid / Invalid']="Not checked"
    #for index,row in df[7:8].iterrows():
    for index,row in df[0:len(df)].iterrows():
        #print(row['Full AEM image'])
        s = requests.Session()
        #s.proxies = proxies
        try:
            print(index,row[URL_column])
             
            if pd.isna(row[URL_column]) or pd.isnull(row[URL_column]) or (row[URL_column] ==""):
                print(index, " missing")
                df['Valid / Invalid'][index]="Missing link"
            else:   
                r = s.get(row[URL_column])
                r.raise_for_status()
                df.loc[df[URL_column]==row[URL_column], 'Valid / Invalid']="Good link"
                print(index, " good")
        except :
            print(index,"not good")
            #print(e)
            df.loc[df[URL_column]==row[URL_column], 'Valid / Invalid']="Not good"
             
         
     
    print(len(df[df['Valid / Invalid'] == "Good link"])) 
    book = load_workbook(files)
    writer = pd.ExcelWriter(files, engine='openpyxl')
    writer.book = book
    df.to_excel(writer,sheet_name="Links checked", index=False, header=True,startrow=0, columns=df.columns)
    df=df.replace('<NA>', '')
    writer.close()
 
elif Menu_selection =="Continue Upload":
    #GUI to select project/project id
    project_id=[]
    project_id=Project_List.loc[Project_List['name']==choicebox("Select the project you want to work on.Warning if duplicate name exist, only first in the list willbe selected"  , "Optimizely Uploader", sorted(Project_List["name"]) ) ,'id'].item()
    ##List projects
    #List audience
    Audience_List=list_audiences(project_id,headers)       
    ###List Audience
    df,files=openExcel()
    Campaigns_List=List_campaigns(project_id)
    Campaigns_List=Campaigns_List.sort_values(by='created', ascending=False)
    Campaign_selection=choicebox("Select the campaign you want to work on"  , "Optimizely Uploader",sorted(Campaigns_List["name"]) )
    Campaign_id=[]
    Campaign_id=Campaigns_List.loc[Campaigns_List['name']==Campaign_selection ,'id'].tolist()[0]
     
    start = time.time()
       
    duplicate_experiments()  
    pull_QA_List()
    Final=upload_experiments(df[~df['Experiment Name'].isin(QA_df['Experiment Name'])])
      
    end = time.time()
    print(end - start)
    print(str(datetime. timedelta(seconds=end - start)))
     
elif Menu_selection =="Update experiments - can only add new ones and prioritise":
    Project_List=list_projects()  
    project_id=[]
    project_id=Project_List.loc[Project_List['name']==choicebox("Select the project you want to work on.Warning if duplicate name exist, only first in the list willbe selected"  , "Optimizely Uploader", sorted(Project_List["name"]) ) ,'id'].item()
       
    Campaigns_List=List_campaigns(project_id)
    Campaigns_List=Campaigns_List.sort_values(by='created', ascending=False)
    Campaign_selection=choicebox("Select the campaign you want to work on"  , "Optimizely Uploader",sorted(Campaigns_List["name"]) )
    Campaign_id=[]
    Campaign_id=Campaigns_List.loc[Campaigns_List['name']==Campaign_selection ,'id'].tolist()[0]
    Pages_List=list_pages(project_id,headers)
    Pages_List=Pages_List[Pages_List['archived'] != True]
    Pages_id=int(Pages_List.loc[Pages_List["Page name"]==choicebox("Select the page where the campaign runs on."  , "Optimizely Uploader",  sorted(Pages_List["Page name"]) ) ,'Page ID'].item())
     
    QA_df=list_experiments(project_id,headers,Campaign_id)
    QA_df=QA_df[QA_df["Experiment status"] !="archived"]
    #Audience_List.rename(columns={'audience_name': "Audience name",'audience_value' : 'Audience Value','audience_id' : "Audience ID",'project_id':"Project ID"}, inplace=True)
    Audience_List=list_audiences(project_id,headers)
    Audience_List=Audience_List[Audience_List["archived"] ==False]
    Pages_List=list_pages(project_id,headers)
    Pages_List=Pages_List[Pages_List["archived"] ==False]
    QA_df=pd.merge(QA_df.astype(str), Audience_List.astype(str), left_on=  ["Project ID",'Audience ID'], right_on= ["Project ID",'Audience ID'], how = 'left', suffixes=('', '_OPTIMIZELY')).filter(regex='^(?!.*_OPTIMIZELY)')
    QA_df.columns
    QA_df=pd.merge(QA_df.astype(str), Pages_List.astype(str), left_on=  ["page_id"], right_on= ['Page ID'], how = 'left', suffixes=('', '_OPTIMIZELY')).filter(regex='^(?!.*_OPTIMIZELY)')
    QA_df.drop('Page ID', axis=1, inplace=True)
    #QA_df.drop_duplicates(subset=["Project ID",'Audience ID'], inplace=True)
    print(QA_df)
    Audience_List.drop_duplicates(subset=["Project ID",'Audience ID'], inplace=True)  
    df,files=openExcel()
    merge_list= template.columns.tolist()
    merge_list.remove('AEM Audience')
    merge_list.remove('Description') 
    merge_list.remove('Audience Description')
    #test=df[df['Experiment Name']=='UK-WPB | Regulatory | ISOL | Premier with core products, PIB active but missing mobile/EM | DC | 528']
    #test=test.append(QA_df[QA_df['Experiment Name']=='UK-WPB | Regulatory | ISOL | Premier with core products, PIB active but missing mobile/EM | DC | 528'])
    QA_done=pd.merge(df.astype(str),QA_df.astype(str),left_on=merge_list, right_on=merge_list , how = 'left',indicator=True)
    QA_done=pd.merge(QA_done.astype(str),QA_df[['Project ID','Experiment ID','Experiment Name']].astype(str),left_on='Experiment Name', right_on='Experiment Name' , how = 'left')
    df=pd.merge(df.astype(str),QA_done[['Experiment Name','_merge','Project ID_y','Experiment ID_y']].astype(str),left_on='Experiment Name', right_on='Experiment Name' , how = 'left', suffixes=('', '_OPTIMIZELY')).filter(regex='^(?!.*_OPTIMIZELY)')
    upload_experiments(df)
    prioritisation()
     
elif Menu_selection =="Download experiment details":
    Project_List=list_projects()  
    project_id=[]
    project_id=Project_List.loc[Project_List['name']==choicebox("Select the project you want to work on.Warning if duplicate name exist, only first in the list willbe selected"  , "Optimizely Uploader", sorted(Project_List["name"]) ) ,'id'].item()
    project_name=Project_List[Project_List['id']==project_id]['name'].tolist()[0] 
    Campaigns_List=List_campaigns(project_id)
    Campaigns_List=Campaigns_List.sort_values(by='created', ascending=False)
    Campaign_selection=choicebox("Select the campaign you want to work on"  , "Optimizely Uploader",sorted(Campaigns_List["name"]) )
    Campaign_id=[]
    Campaign_id=Campaigns_List.loc[Campaigns_List['name']==Campaign_selection ,'id'].tolist()[0]
    Experiments_df=list_experiments(project_id,headers,Campaign_id)
     
    folder=diropenbox('Select location for file')
    path = folder+'\\'+project_name+'_' + Campaign_selection + '_Experiment Details_'+re.sub(":","-",str(datetime.datetime.now()) )+'.xlsx'   
    writer = pd.ExcelWriter(path, engine='openpyxl')
    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
    Experiments_df.to_excel(writer,sheet_name="Experiment details", index=False, header=True,startrow=0, columns=Experiments_df.columns)
    writer.close()
 
 
elif Menu_selection =="Check for External links":######this is not working, we need to be able to pull all type of test not just personalisation
    #The point of this is to check if there is any url that have been implemented and would redirect outside the bank network to prevent malicious redirect
    #GUI to select project/project id
    project_id=[]
    project_id=Project_List.loc[Project_List['name']==choicebox("Select the project you want to work on.Warning if duplicate name exist, only first in the list willbe selected"  , "Optimizely Uploader", sorted(Project_List["name"]) ) ,'id'].item()
    ##List projects
    #Campaigns_List=List_campaigns(project_id)
    #Campaigns_List=Campaigns_List.sort_values(by='created', ascending=False)
    Experiments_lict_complete=[]
    for P in Campaigns:
        if 'archived' not in P['status']:
            print(P['status'])
            Experiments_lict_complete.append({
                "name" : P["name"],
               'experiment_priorities': P.get("experiment_priorities","")
     
            })
             
    Experiments_lict_complete=pd.DataFrame(Experiments_lict_complete)
    Experiments_lict_complete=Experiments_lict_complete[Experiments_lict_complete['experiment_priorities']!=""]
    url="https://api.optimizely.com/v2/experiments/"
    params = {
    "campaign_id" : Campaign_id,
    "project_id" : project_id
    }   
    r_total_json=[]
    from urlextract import URLExtract
    extractor = URLExtract()
    for i in Experiments_lict_complete['name']:
        #print(i)
        for j in  Experiments_lict_complete['experiment_priorities']:
            #print(j)
            for k in j:
                print(url+str(k[0]))
                r = requests.get(url+str(k[0]), params=params, headers=headers)
                urls = extractor.find_urls(r.text)
                Experiments_containing_external_domain=[]
                for P in urls:
                        Experiments_containing_external_domain.append({
                            "name" : i,
                           'URL_found':urls
                 
                        })
                         
    Experiments_containing_external_domain=pd.DataFrame(Experiments_containing_external_domain)
